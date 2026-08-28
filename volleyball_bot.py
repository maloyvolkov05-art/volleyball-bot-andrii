"""
Волейбольний бот з обліком відвідування та оплати
- Щочетверга о 12:00 надсилає ОДНЕ опитування на найближчу суботу (дублі заблоковані)
- Хто натиснув ✅ → автоматично в Excel як "прийде"
- /attended — тільки для людей поза Telegram
- /paid — тільки в особистому чаті з ботом
- /excel — Excel таблиця
"""

import asyncio
import json
import logging
import os
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, PollAnswer
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from aiohttp import web, ClientSession

logging.basicConfig(level=logging.INFO)

# ============================================================
import os

BOT_TOKEN = os.environ.get("BOT_TOKEN", "YOUR_BOT_TOKEN")
CHAT_ID   = int(os.environ.get("CHAT_ID", 0)) or None
PRICE     = 2.5
DATA_FILE = "volleyball_data.json"
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL   = "gemini-3.6-flash"  # безкоштовна модель; якщо Google її зніме з підтримки — заміните на актуальну
# ============================================================

bot = Bot(token=BOT_TOKEN)
dp  = Dispatcher()
scheduler = AsyncIOScheduler(timezone="Europe/Kyiv")


# ============================================================
# JSON база
# ============================================================
def load_data() -> dict:
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"sessions": {}}


def save_data(data: dict) -> None:
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def get_next_saturday() -> str:
    """Повертає найближчу суботу (включно з сьогодні, якщо сьогодні субота)."""
    today = datetime.now()
    days_ahead = (5 - today.weekday()) % 7  # 0 якщо сьогодні субота
    return (today + timedelta(days=days_ahead)).strftime("%Y-%m-%d")


def ensure_session(data: dict, saturday: str) -> None:
    if saturday not in data["sessions"]:
        data["sessions"][saturday] = {"poll_id": None, "players": {}}


def find_player(session: dict, search: str) -> tuple[str | None, dict | None]:
    """Шукає гравця за username або іменем. Повертає (uid, player)."""
    search = search.lower()
    for uid, p in session.get("players", {}).items():
        if (p.get("username", "").lower() == search or
                p.get("name", "").lower() == search):
            return uid, p
    return None, None


# ============================================================
# Команди
# ============================================================

@dp.message(Command("start"))
async def cmd_start(message: Message):
    await message.answer(
        "🏐 <b>Волейбольний бот</b>\n\n"
        "Щочетверга о 12:00 надсилаю одне опитування на найближчу суботу.\n"
        "Хто натискає ✅ — автоматично відмічається як учасник.\n\n"
        "<b>Команди:</b>\n"
        "/status   — хто йде, скільки зібрано\n"
        "/next     — дата наступної гри\n"
        "/summary  — підсумок по всіх іграх\n"
        "/excel    — таблиця Excel\n"
        "/poll     — надіслати опитування зараз\n"
        "/id       — ID цього чату\n\n"
        "<b>Для людей поза групою:</b>\n"
        "/attended ім'я — вручну додати людину",
        parse_mode="HTML"
    )


@dp.message(Command("id"))
async def cmd_id(message: Message):
    await message.answer(f"ID цього чату: <code>{message.chat.id}</code>", parse_mode="HTML")


@dp.message(Command("next"))
async def cmd_next(message: Message):
    saturday = get_next_saturday()
    dt = datetime.strptime(saturday, "%Y-%m-%d")
    await message.answer(
        f"📅 Наступна гра: <b>субота {dt.strftime('%d.%m.%Y')} о 12:00</b>",
        parse_mode="HTML"
    )


# ============================================================
# Опитування
# ============================================================

async def send_poll(chat_id: int = None):
    target = chat_id or CHAT_ID
    if not target:
        logging.warning("CHAT_ID не вказано!")
        return

    saturday = get_next_saturday()
    data = load_data()
    ensure_session(data, saturday)

    # ── Захист від дублів: якщо опитування вже є — не надсилаємо ──
    if data["sessions"][saturday].get("poll_id"):
        logging.info(f"Опитування на {saturday} вже існує — пропускаємо.")
        return

    dt = datetime.strptime(saturday, "%Y-%m-%d")

    msg = await bot.send_poll(
        chat_id=target,
        question=f"🏐 Волейбол у суботу {dt.strftime('%d.%m')} о 12:00 — йдеш?",
        options=["✅ Так, буду!", "❌ Не зможу"],
        is_anonymous=False,
        allows_multiple_answers=False
    )

    data["sessions"][saturday]["poll_id"] = str(msg.poll.id)
    save_data(data)
    logging.info(f"Опитування надіслано для {saturday}")


@dp.message(Command("poll"))
async def cmd_poll(message: Message):
    saturday = get_next_saturday()
    data = load_data()
    if data.get("sessions", {}).get(saturday, {}).get("poll_id"):
        dt = datetime.strptime(saturday, "%Y-%m-%d")
        await message.answer(
            f"⚠️ Опитування на суботу {dt.strftime('%d.%m.%Y')} вже надіслано.\n"
            "Нове не створюється — щоб уникнути дублів."
        )
        return
    await send_poll(message.chat.id)


# ============================================================
# Відповідь на опитування → автоматично в Excel
# ============================================================

@dp.poll_answer()
async def handle_poll_answer(poll_answer: PollAnswer):
    data = load_data()
    user   = poll_answer.user
    option = poll_answer.option_ids[0]  # 0=Так, 1=Ні
    voted_yes = (option == 0)

    # Знаходимо сесію
    session_key = None
    for date, session in data["sessions"].items():
        if session.get("poll_id") == str(poll_answer.poll_id):
            session_key = date
            break

    if not session_key:
        return

    uid  = str(user.id)
    name = f"{user.first_name or ''} {user.last_name or ''}".strip()

    players = data["sessions"][session_key]["players"]

    if uid not in players:
        players[uid] = {
            "name": name,
            "username": user.username or "",
            "attended": False,
            "paid": False
        }

    players[uid]["name"]     = name
    players[uid]["username"] = user.username or ""

    # ✅ Так → attended = True автоматично
    # ❌ Ні  → attended = False
    players[uid]["attended"] = voted_yes

    save_data(data)
    logging.info(f"{name} → {'✅ буде' if voted_yes else '❌ не буде'} ({session_key})")


# ============================================================
# /attended — вручну для людей поза Telegram
# ============================================================

@dp.message(Command("attended"))
async def cmd_attended(message: Message):
    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.answer("Використання: /attended Ім'я Прізвище")
        return

    name_raw = args[1].strip()
    saturday = get_next_saturday()
    data     = load_data()
    ensure_session(data, saturday)

    players = data["sessions"][saturday]["players"]

    # Перевіряємо чи вже є такий
    uid, existing = find_player(data["sessions"][saturday], name_raw.lstrip("@"))
    if existing:
        existing["attended"] = True
        await message.answer(f"✅ {name_raw} вже є в списку — відмічено як учасник.")
    else:
        # Новий гравець (поза Telegram) — генеруємо uid на основі імені
        fake_uid = f"manual_{name_raw.lower().replace(' ', '_')}"
        players[fake_uid] = {
            "name": name_raw,
            "username": "",
            "attended": True,
            "paid": False
        }
        await message.answer(f"✅ {name_raw} додано як учасник суботи.")

    save_data(data)


# ============================================================
# /paid — ТІЛЬКИ в особистому чаті з ботом
# ============================================================

@dp.message(Command("paid"))
async def cmd_paid(message: Message):
    # Перевіряємо що це особистий чат (не група)
    if message.chat.type != "private":
        # Мовчки видаляємо повідомлення в групі
        try:
            await message.delete()
        except Exception:
            pass
        return

    args = message.text.split(maxsplit=1)
    if len(args) < 2:
        await message.answer(
            "Використання: /paid @username або /paid Ім'я\n\n"
            "Приклад: /paid @andriy або /paid Андрій"
        )
        return

    search   = args[1].strip().lstrip("@")
    saturday = get_next_saturday()
    data     = load_data()

    if saturday not in data["sessions"]:
        await message.answer("Немає активної сесії на цю суботу.")
        return

    uid, player = find_player(data["sessions"][saturday], search)

    if not player:
        await message.answer(f"❌ Гравця «{args[1]}» не знайдено в списку на цю суботу.")
        return

    data["sessions"][saturday]["players"][uid]["paid"] = True
    save_data(data)

    name = f"@{player['username']}" if player.get("username") else player["name"]
    dt   = datetime.strptime(saturday, "%Y-%m-%d")
    await message.answer(
        f"✅ {name} — оплату відмічено (+{PRICE}€)\n"
        f"📅 Субота {dt.strftime('%d.%m.%Y')}"
    )


# ============================================================
# /status
# ============================================================

@dp.message(Command("status"))
async def cmd_status(message: Message):
    data     = load_data()
    saturday = get_next_saturday()

    if saturday not in data["sessions"]:
        await message.answer("📭 Опитування ще не проводилось для цієї суботи.")
        return

    players   = data["sessions"][saturday].get("players", {})
    attending = [p for p in players.values() if p.get("attended")]
    paid      = [p for p in attending if p.get("paid")]
    unpaid    = [p for p in attending if not p.get("paid")]

    collected = len(paid) * PRICE
    needed    = len(attending) * PRICE
    dt        = datetime.strptime(saturday, "%Y-%m-%d")

    text  = f"🏐 <b>Субота {dt.strftime('%d.%m.%Y')}</b>\n\n"
    text += f"👥 Йдуть: <b>{len(attending)}</b> людей\n"
    text += f"💶 Потрібно: <b>{needed:.1f}€</b>\n"
    text += f"✅ Зібрано: <b>{collected:.1f}€</b>\n"
    text += f"⏳ Залишилось: <b>{needed - collected:.1f}€</b>\n"

    if paid:
        text += "\n💚 <b>Заплатили:</b>\n"
        for p in paid:
            n = f"@{p['username']}" if p.get("username") else p["name"]
            text += f"  • {n}\n"

    if unpaid:
        text += "\n🔴 <b>Не заплатили:</b>\n"
        for p in unpaid:
            n = f"@{p['username']}" if p.get("username") else p["name"]
            text += f"  • {n} (−{PRICE}€)\n"

    await message.answer(text, parse_mode="HTML")


# ============================================================
# /summary
# ============================================================

@dp.message(Command("summary"))
async def cmd_summary(message: Message):
    data     = load_data()
    sessions = data.get("sessions", {})

    if not sessions:
        await message.answer("Даних ще немає.")
        return

    text = "📊 <b>Загальний підсумок</b>\n\n"
    total_needed = total_paid = 0.0

    for date, session in sorted(sessions.items()):
        players    = session.get("players", {})
        attending  = [p for p in players.values() if p.get("attended")]
        paid_count = sum(1 for p in attending if p.get("paid"))
        needed     = len(attending) * PRICE
        collected  = paid_count * PRICE
        total_needed += needed
        total_paid   += collected
        dt = datetime.strptime(date, "%Y-%m-%d")
        text += f"📅 <b>{dt.strftime('%d.%m.%Y')}</b> — {len(attending)} людей | {collected:.1f}/{needed:.1f}€\n"

    text += f"\n💶 <b>Всього потрібно:</b> {total_needed:.1f}€"
    text += f"\n✅ <b>Всього зібрано:</b> {total_paid:.1f}€"
    text += f"\n⏳ <b>Борг:</b> {total_needed - total_paid:.1f}€"
    await message.answer(text, parse_mode="HTML")


# ============================================================
# /excel
# ============================================================

@dp.message(Command("excel"))
async def cmd_excel(message: Message):
    data     = load_data()
    sessions = data.get("sessions", {})

    if not sessions:
        await message.answer("Даних ще немає.")
        return

    try:
        filename = build_excel(sessions)
        with open(filename, "rb") as f:
            file_bytes = f.read()
        from aiogram.types import BufferedInputFile
        doc = BufferedInputFile(file_bytes, filename="volleyball.xlsx")
        await message.answer_document(document=doc, caption="📊 Таблиця відвідування та оплати")
    except Exception as e:
        logging.exception("Помилка при створенні Excel")
        await message.answer(f"❌ Помилка при створенні таблиці: {e}")


def build_excel(sessions: dict, filename: str = "volleyball.xlsx") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Відвідування"

    header_fill  = PatternFill("solid", fgColor="1F5C99")
    paid_fill    = PatternFill("solid", fgColor="C6EFCE")
    unpaid_fill  = PatternFill("solid", fgColor="FFCCCC")
    absent_fill  = PatternFill("solid", fgColor="EEEEEE")
    neutral_fill = PatternFill("solid", fgColor="F2F2F2")
    total_fill   = PatternFill("solid", fgColor="D9E1F2")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    bold_font    = Font(bold=True)
    center       = Alignment(horizontal="center", vertical="center")
    thin         = Side(style="thin", color="CCCCCC")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Всі унікальні гравці
    all_players: dict[str, str] = {}
    for session in sessions.values():
        for uid, p in session.get("players", {}).items():
            display = f"@{p['username']}" if p.get("username") else p.get("name", uid)
            all_players[uid] = display

    sorted_dates = sorted(sessions.keys())

    # Заголовки
    ws.cell(1, 1, "Гравець").font = header_font
    ws.cell(1, 1).fill      = header_fill
    ws.cell(1, 1).alignment = center
    ws.cell(1, 1).border    = border
    ws.column_dimensions["A"].width = 22

    for col, date in enumerate(sorted_dates, start=2):
        dt   = datetime.strptime(date, "%Y-%m-%d")
        cell = ws.cell(1, col, dt.strftime("%d.%m.%Y"))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    sc = len(sorted_dates) + 2
    for i, label in enumerate(["Ігор", "Сплачено (€)", "Борг (€)"]):
        cell = ws.cell(1, sc + i, label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(sc + i)].width = 14

    # Рядки гравців
    for row, (uid, display_name) in enumerate(sorted(all_players.items(), key=lambda x: x[1]), start=2):
        ws.cell(row, 1, display_name).fill = neutral_fill
        ws.cell(row, 1).border    = border
        ws.cell(row, 1).alignment = Alignment(vertical="center")

        games = paid_sum = 0

        for col, date in enumerate(sorted_dates, start=2):
            session = sessions.get(date, {})
            player  = session.get("players", {}).get(uid)

            if not player or not player.get("attended"):
                cell = ws.cell(row, col, "—")
                cell.fill = absent_fill
            else:
                games += 1
                if player.get("paid"):
                    paid_sum += 1
                    cell = ws.cell(row, col, "✅ Сплатив")
                    cell.fill = paid_fill
                else:
                    cell = ws.cell(row, col, "⏳ Борг")
                    cell.fill = unpaid_fill

            cell.alignment = center
            cell.border    = border

        debt = (games - paid_sum) * PRICE
        for i, val in enumerate([games, f"{paid_sum * PRICE:.1f}", f"{debt:.1f}"]):
            c = ws.cell(row, sc + i, val)
            c.alignment = center
            c.border     = border
            c.fill       = neutral_fill

    # Підсумковий рядок
    total_row = len(all_players) + 2
    ws.cell(total_row, 1, "РАЗОМ").font = bold_font
    ws.cell(total_row, 1).fill   = total_fill
    ws.cell(total_row, 1).border = border

    for col, date in enumerate(sorted_dates, start=2):
        session   = sessions.get(date, {})
        players   = session.get("players", {})
        attending = sum(1 for p in players.values() if p.get("attended"))
        paid_cnt  = sum(1 for p in players.values() if p.get("attended") and p.get("paid"))
        cell = ws.cell(total_row, col, f"{paid_cnt * PRICE:.1f}/{attending * PRICE:.1f}€")
        cell.font      = bold_font
        cell.fill      = total_fill
        cell.alignment = center
        cell.border    = border

    ws.freeze_panes = "B2"
    wb.save(filename)
    return filename


# ============================================================
# Нейронка (Google Gemini) — відповідає на питання в чаті
# ============================================================

BOT_USERNAME = None  # заповнюється при старті


async def ask_gemini(question: str) -> str:
    """Надсилає запит до Gemini API і повертає текст відповіді."""
    if not GEMINI_API_KEY:
        return "⚠️ Нейронка ще не підключена — не заданий GEMINI_API_KEY."

    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/"
        f"{GEMINI_MODEL}:generateContent"
    )

    async with ClientSession() as session:
        try:
            async with session.post(
                url,
                headers={
                    "x-goog-api-key": GEMINI_API_KEY,
                    "content-type": "application/json",
                },
                json={
                    "system_instruction": {
                        "parts": [{
                            "text": (
                                "Ти — дружній помічник волейбольного чату. "
                                "Відповідай коротко (2-4 речення), українською, "
                                "по суті питання."
                            )
                        }]
                    },
                    "contents": [
                        {"role": "user", "parts": [{"text": question}]}
                    ],
                },
                timeout=30,
            ) as resp:
                data = await resp.json()
                if resp.status != 200:
                    logging.error(f"Gemini API error: {data}")
                    return "❌ Не вдалося отримати відповідь від нейронки."
                candidates = data.get("candidates", [])
                if not candidates:
                    return "🤔 Нейронка не дала відповіді."
                parts = candidates[0].get("content", {}).get("parts", [])
                text = "".join(p.get("text", "") for p in parts).strip()
                return text or "🤔 Нейронка не дала відповіді."
        except Exception:
            logging.exception("Помилка запиту до Gemini")
            return "❌ Сталася помилка при зверненні до нейронки."


def _is_addressed_to_bot(message: Message) -> bool:
    """У групі бот відповідає тільки якщо його згадали (@ім'я) або
    відповіли на його повідомлення. В особистому чаті — завжди."""
    if message.chat.type == "private":
        return True
    if message.reply_to_message and message.reply_to_message.from_user and \
            message.reply_to_message.from_user.is_bot and \
            message.reply_to_message.from_user.username == BOT_USERNAME:
        return True
    if BOT_USERNAME and message.text and f"@{BOT_USERNAME}" in message.text:
        return True
    return False


@dp.message(F.text & ~F.text.startswith("/"))
async def handle_ai_question(message: Message):
    if not _is_addressed_to_bot(message):
        return

    question = message.text
    if BOT_USERNAME:
        question = question.replace(f"@{BOT_USERNAME}", "").strip()

    if not question:
        return

    # ── Жартівлива фіксована відповідь ──
    q_lower = question.lower().strip(" ?!.")
    if q_lower in ("хто твій тато", "хто твій батько", "хто твой папа", "кто твой папа"):
        await message.reply("Андрій Волков 🏐")
        return

    await bot.send_chat_action(message.chat.id, "typing")
    answer = await ask_gemini(question)
    await message.reply(answer)


# ============================================================
# Планувальник
# ============================================================
def setup_scheduler():
    scheduler.add_job(
        send_poll,
        trigger="cron",
        day_of_week="thu",
        hour=12,
        minute=0,
        timezone="Europe/Kyiv"
    )
    scheduler.start()
    logging.info("Планувальник запущено — щочетверга о 12:00")


async def handle_health(request):
    return web.Response(text="ok")


async def run_health_server():
    """Мінімальний HTTP-сервер — потрібен тільки щоб Render (Web Service)
    бачив відкритий порт і не вважав деплой невдалим."""
    port = int(os.environ.get("PORT", 8080))
    app = web.Application()
    app.router.add_get("/", handle_health)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", port)
    await site.start()
    logging.info(f"Health-check сервер запущено на порту {port}")


async def main():
    global BOT_USERNAME
    me = await bot.get_me()
    BOT_USERNAME = me.username
    logging.info(f"Бот запущено як @{BOT_USERNAME}")

    setup_scheduler()
    await run_health_server()
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())
